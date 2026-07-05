"""
specimen/_analyze_all.py
=========================
각 specimen 파일을 읽어 영역별 전처리 방법을 분석하고,
specimen/result/ 에 {filename}.csv (정제본) + {filename}.txt (분석 로그) 저장.

분석 항목:
- 파일 메타 (크기, 인코딩 추정, 구분자 추정)
- 구조 분석 (헤더 위치, 데이터 영역, 푸터/주석 행, 사이드 주석 열)
- 셀 품질 (결측률, 타입 혼합, 이상값, CSV Injection, 중복행)
- 영역별 전처리 권장사항
"""

from __future__ import annotations

import csv
import io
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

SPECIMEN_DIR = Path(__file__).resolve().parent
RESULT_DIR = SPECIMEN_DIR / "result"

# ──────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────

MISSING_MARKERS = {"", "N/A", "n/a", "NA", "na", "NULL", "null", "NaN", "nan",
                   "None", "none", "-", "--", "?", "  ", "미입력", "#N/A", "#REF!",
                   ".", ".."}


def guess_delimiter(text: str) -> str:
    """첫 5줄로 구분자를 추정."""
    lines = text.strip().split("\n")[:5]
    candidates = {",": 0, ";": 0, "\t": 0, "|": 0}
    for line in lines:
        for sep in candidates:
            candidates[sep] += line.count(sep)
    best = max(candidates, key=candidates.get)
    return best if candidates[best] > 0 else ","


def is_numeric(val: str) -> bool:
    """숫자 파싱 가능 여부."""
    val = val.strip().replace(",", "").replace("₩", "").replace("$", "")
    val = val.replace("원", "").replace("만원", "").replace("USD", "")
    val = val.strip("() '\"")
    if not val:
        return False
    try:
        float(val)
        return True
    except ValueError:
        return False


def is_date_like(val: str) -> bool:
    """날짜 패턴 감지 (간이)."""
    date_patterns = [
        r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}",
        r"\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}",
        r"\d{1,2}\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)",
        r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2}",
        r"\d{8}",
    ]
    for pat in date_patterns:
        if re.search(pat, val, re.IGNORECASE):
            return True
    return False


def strip_sanitize_prefix(val: str) -> str:
    """CSV sanitization 접두어(') 제거."""
    if val.startswith("'") and len(val) > 1:
        rest = val[1:]
        if rest and rest[0] in ("=", "+", "-", "@"):
            return rest
    return val


def clean_cell(val: str) -> str:
    """기본 셀 정제."""
    if val is None:
        return ""
    val = str(val).strip()
    val = val.replace("\x00", "")  # null byte
    val = strip_sanitize_prefix(val)
    if val in MISSING_MARKERS:
        return ""
    return val


# ──────────────────────────────────────────
# CSV / Text file reader
# ──────────────────────────────────────────

def read_csv_file(path: Path) -> tuple[list[list[str]], dict]:
    """CSV를 읽어 2D 리스트 + 메타 반환."""
    raw = path.read_bytes()
    meta = {"size_bytes": len(raw), "format": "CSV"}

    # Encoding 감지
    for enc in ["utf-8-sig", "utf-8", "cp949", "euc-kr", "latin-1"]:
        try:
            text = raw.decode(enc)
            meta["encoding"] = enc
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        text = raw.decode("latin-1", errors="replace")
        meta["encoding"] = "latin-1 (fallback)"

    # Delimiter
    delim = guess_delimiter(text)
    meta["delimiter"] = repr(delim)

    # Parse
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = [row for row in reader]
    meta["total_rows"] = len(rows)
    meta["max_cols"] = max((len(r) for r in rows), default=0)
    meta["min_cols"] = min((len(r) for r in rows if r), default=0)

    return rows, meta


def read_xlsx_file(path: Path) -> list[tuple[str, list[list], dict]]:
    """XLSX의 각 시트를 읽어 [(시트명, rows, meta), ...] 반환."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
        meta = {
            "format": "XLSX",
            "sheet_name": name,
            "total_rows": len(rows),
            "max_cols": max((len(r) for r in rows), default=0),
            "size_bytes": path.stat().st_size,
        }
        sheets.append((name, rows, meta))
    wb.close()
    return sheets


# ──────────────────────────────────────────
# Analysis Engine
# ──────────────────────────────────────────

class TableAnalyzer:
    """테이블 구조 및 품질 분석."""

    def __init__(self, rows: list[list[str]], meta: dict):
        self.rows = rows
        self.meta = meta
        self.log_lines: list[str] = []
        self.issues: list[str] = []
        self.recommendations: list[str] = []

        # Analysis results
        self.header_row_idx: Optional[int] = None
        self.data_start_idx: Optional[int] = None
        self.data_end_idx: Optional[int] = None
        self.num_data_cols: int = 0
        self.empty_rows: list[int] = []
        self.footer_rows: list[int] = []
        self.annotation_rows: list[int] = []
        self.annotation_cols: list[int] = []
        self.duplicate_rows: list[int] = []

    def log(self, msg: str):
        self.log_lines.append(msg)

    def analyze(self):
        self.log("=" * 70)
        self.log(f"  분석 시작: {self.meta.get('sheet_name', self.meta.get('format', ''))}")
        self.log(f"  {self.meta}")
        self.log("=" * 70)

        self._find_structure()
        self._analyze_columns()
        self._find_duplicates()
        self._find_annotations()
        self._check_injection()
        self._summarize()

    def _find_structure(self):
        """헤더, 데이터 영역, 빈 행, 푸터를 찾는다."""
        self.log("\n── 구조 분석 ──")

        if not self.rows:
            self.log("  빈 테이블")
            return

        # 빈 행 탐지
        for i, row in enumerate(self.rows):
            if all(clean_cell(c) == "" for c in row):
                self.empty_rows.append(i)

        # 헤더 행 탐지 (첫 번째로 대부분 셀이 비어있지 않고, 숫자가 아닌 행)
        for i, row in enumerate(self.rows):
            cleaned = [clean_cell(c) for c in row]
            non_empty = [c for c in cleaned if c]
            if len(non_empty) < 2:
                continue
            numeric_ratio = sum(1 for c in non_empty if is_numeric(c)) / max(len(non_empty), 1)
            if numeric_ratio < 0.5:  # 절반 이상이 텍스트면 헤더 후보
                self.header_row_idx = i
                self.num_data_cols = len(row)
                break

        if self.header_row_idx is None:
            self.header_row_idx = 0
            self.num_data_cols = len(self.rows[0]) if self.rows else 0

        # 데이터 시작 (헤더 바로 다음 비빈행)
        self.data_start_idx = self.header_row_idx + 1
        while self.data_start_idx < len(self.rows) and self.data_start_idx in self.empty_rows:
            self.data_start_idx += 1

        # 데이터 끝 + 푸터 탐지 (하단부에서 연속된 비데이터 행)
        self.data_end_idx = len(self.rows) - 1
        for i in range(len(self.rows) - 1, self.data_start_idx - 1, -1):
            row = self.rows[i]
            cleaned = [clean_cell(c) for c in row]
            non_empty = [c for c in cleaned if c]
            if len(non_empty) == 0:
                self.data_end_idx = i - 1
                continue
            # 한두 셀만 있고 나머지 빈 → 푸터/주석
            if len(non_empty) <= 2 and len(row) > 3:
                self.footer_rows.append(i)
                self.data_end_idx = i - 1
            else:
                break

        self.log(f"  빈 행: {len(self.empty_rows)}개 (idx: {self.empty_rows[:10]}{'...' if len(self.empty_rows) > 10 else ''})")
        self.log(f"  헤더 행: {self.header_row_idx}")
        self.log(f"  데이터 영역: row {self.data_start_idx} ~ {self.data_end_idx}")
        self.log(f"  데이터 열 수: {self.num_data_cols}")
        self.log(f"  푸터 행: {self.footer_rows}")

        if self.header_row_idx > 0:
            self.annotation_rows.extend(range(0, self.header_row_idx))
            self.issues.append(f"헤더 위 메타데이터/제목 행 {self.header_row_idx}개")
            self.recommendations.append(f"ROW 0~{self.header_row_idx - 1}: 제목/메타데이터 행 제거")

        if self.empty_rows:
            self.issues.append(f"빈 행 {len(self.empty_rows)}개 산재")
            self.recommendations.append(f"빈 행 {len(self.empty_rows)}개 제거")

        if self.footer_rows:
            self.issues.append(f"하단 푸터/주석 행 {len(self.footer_rows)}개")
            self.recommendations.append(f"ROW {min(self.footer_rows)}~: 하단 푸터/주석 행 제거")

    def _analyze_columns(self):
        """열별 타입, 결측률, 이상값 분석."""
        self.log("\n── 열별 분석 ──")

        if self.data_start_idx is None or self.data_end_idx is None:
            return
        if self.data_start_idx > self.data_end_idx:
            return

        data_rows = [self.rows[i] for i in range(self.data_start_idx, min(self.data_end_idx + 1, len(self.rows)))
                      if i not in self.empty_rows and i not in self.footer_rows]

        if not data_rows:
            return

        headers = self.rows[self.header_row_idx] if self.header_row_idx is not None else []
        ncols = max(len(r) for r in data_rows) if data_rows else 0

        col_stats = []
        for col_idx in range(ncols):
            header_name = headers[col_idx].strip() if col_idx < len(headers) else f"Col_{col_idx}"
            header_name = clean_cell(header_name) or f"Col_{col_idx}"

            values = []
            for row in data_rows:
                val = clean_cell(row[col_idx]) if col_idx < len(row) else ""
                values.append(val)

            total = len(values)
            missing = sum(1 for v in values if v == "")
            non_empty = [v for v in values if v]

            # 타입 추론
            type_counts = Counter()
            for v in non_empty:
                if is_numeric(v):
                    type_counts["numeric"] += 1
                elif is_date_like(v):
                    type_counts["date"] += 1
                else:
                    type_counts["text"] += 1

            dominant_type = type_counts.most_common(1)[0][0] if type_counts else "empty"
            mixed = len(type_counts) > 1

            missing_pct = missing / max(total, 1) * 100

            stat = {
                "name": header_name,
                "missing": missing,
                "missing_pct": missing_pct,
                "types": dict(type_counts),
                "dominant": dominant_type,
                "mixed": mixed,
                "total": total,
                "unique": len(set(non_empty)),
            }
            col_stats.append(stat)

            self.log(f"  [{col_idx}] {header_name:20s} | 결측: {missing_pct:5.1f}% | 타입: {dict(type_counts)} | unique: {stat['unique']}")

            if missing_pct > 50:
                self.issues.append(f"열 '{header_name}': 결측률 {missing_pct:.1f}% (과반)")
                self.recommendations.append(f"열 '{header_name}': 결측률 {missing_pct:.0f}% → 열 삭제 또는 보간 검토")
            elif missing_pct > 10:
                self.recommendations.append(f"열 '{header_name}': 결측 {missing_pct:.0f}% → 적절한 결측 처리 필요")

            if mixed:
                self.issues.append(f"열 '{header_name}': 타입 혼합 {dict(type_counts)}")
                self.recommendations.append(f"열 '{header_name}': 타입 통일 필요 (주 타입: {dominant_type})")

            # 빈 열 감지 (사이드 주석 후보)
            if missing_pct > 90:
                self.annotation_cols.append(col_idx)

        if self.annotation_cols:
            self.issues.append(f"거의 빈 열(90%+ 결측): {self.annotation_cols}")
            self.recommendations.append(f"열 {self.annotation_cols}: 사이드 주석 또는 불필요 열 → 삭제 검토")

        # Ragged columns
        col_counts = Counter(len(r) for r in data_rows)
        if len(col_counts) > 1:
            self.issues.append(f"행별 열 수 불일치 (ragged): {dict(col_counts)}")
            mode_cols = col_counts.most_common(1)[0][0]
            self.recommendations.append(f"열 수 불일치 → 주 열 수 {mode_cols}에 맞춰 패딩/트리밍")

    def _find_duplicates(self):
        """중복 행 탐지."""
        self.log("\n── 중복 행 ──")
        if not self.rows:
            return

        seen = {}
        for i, row in enumerate(self.rows):
            key = tuple(clean_cell(c) for c in row)
            if all(c == "" for c in key):
                continue
            if key in seen:
                self.duplicate_rows.append(i)
            else:
                seen[key] = i

        if self.duplicate_rows:
            self.log(f"  중복 행 {len(self.duplicate_rows)}개 발견")
            self.issues.append(f"중복 행 {len(self.duplicate_rows)}개")
            self.recommendations.append(f"중복 행 {len(self.duplicate_rows)}개 제거")
        else:
            self.log("  중복 행 없음")

    def _find_annotations(self):
        """합계/소계 행, 주석 행 탐지."""
        self.log("\n── 합계/소계/주석 행 ──")
        summary_keywords = {"합계", "총합계", "소계", "subtotal", "total", "grand total",
                            "sum", "average", "평균", "※", "주)", "자료:", "작성", "단위:",
                            "생성일", "report generated", "tax"}

        summary_rows = []
        for i, row in enumerate(self.rows):
            text = " ".join(clean_cell(c).lower() for c in row)
            for kw in summary_keywords:
                if kw.lower() in text:
                    summary_rows.append(i)
                    break

        if summary_rows:
            self.log(f"  합계/소계/주석 행: {summary_rows}")
            self.issues.append(f"합계/소계/주석 행 {len(summary_rows)}개")
            self.recommendations.append(f"ROW {summary_rows}: 합계/소계/주석 행 제거")
        else:
            self.log("  합계/소계/주석 행 없음")

        self.summary_rows = summary_rows

    def _check_injection(self):
        """CSV Injection 잔존 패턴 확인."""
        injection_count = 0
        for row in self.rows:
            for cell in row:
                val = str(cell).strip()
                if val and val[0] in ("=", "@") and not is_numeric(val):
                    injection_count += 1
        if injection_count > 0:
            self.issues.append(f"CSV Injection 위험 셀 {injection_count}개 (=, @ 시작)")
            self.recommendations.append(f"CSV Injection 위험 셀 {injection_count}개 → 접두어 제거 또는 무력화")

    def _summarize(self):
        """최종 요약."""
        self.log("\n── 발견된 문제점 ──")
        for i, issue in enumerate(self.issues, 1):
            self.log(f"  [{i}] {issue}")

        self.log("\n── 전처리 권장사항 ──")
        for i, rec in enumerate(self.recommendations, 1):
            self.log(f"  [{i}] {rec}")

    def get_cleaned_rows(self) -> list[list[str]]:
        """분석 결과를 기반으로 정제된 행 반환."""
        if not self.rows:
            return []

        # 제거 대상 행
        skip_rows = set(self.empty_rows) | set(self.footer_rows) | set(self.annotation_rows)
        skip_rows |= set(self.duplicate_rows)
        skip_rows |= set(getattr(self, "summary_rows", []))

        # 헤더
        header_idx = self.header_row_idx if self.header_row_idx is not None else 0
        headers = self.rows[header_idx] if header_idx < len(self.rows) else []
        cleaned_headers = [clean_cell(h) or f"Col_{i}" for i, h in enumerate(headers)]

        # 정제 대상 열 (사이드 주석 열 제거)
        skip_cols = set(self.annotation_cols)

        result = []

        # 헤더 행 추가 (빈 열 제거)
        hdr = [h for i, h in enumerate(cleaned_headers) if i not in skip_cols]
        result.append(hdr)
        target_cols = len(hdr)

        # 데이터 행
        for i in range(header_idx + 1, len(self.rows)):
            if i in skip_rows:
                continue
            row = self.rows[i]
            cleaned = [clean_cell(row[j]) if j < len(row) else "" for j in range(len(headers)) if j not in skip_cols]

            # 빈 행 스킵
            if all(c == "" for c in cleaned):
                continue

            # 열 수 맞추기
            if len(cleaned) < target_cols:
                cleaned.extend([""] * (target_cols - len(cleaned)))
            elif len(cleaned) > target_cols:
                cleaned = cleaned[:target_cols]

            result.append(cleaned)

        return result


# ──────────────────────────────────────────
# Process a single file
# ──────────────────────────────────────────

def process_file(path: Path):
    """파일 하나 분석 + 정제 + 저장."""
    stem = path.stem
    ext = path.suffix.lower()

    print(f"\n{'━' * 60}")
    print(f"  처리 중: {path.name}")
    print(f"{'━' * 60}")

    all_log_lines = []
    all_log_lines.append(f"{'=' * 70}")
    all_log_lines.append(f"파일: {path.name}")
    all_log_lines.append(f"크기: {path.stat().st_size:,} bytes")
    all_log_lines.append(f"분석 시각: {datetime.now().isoformat()}")
    all_log_lines.append(f"{'=' * 70}\n")

    combined_clean_rows = []

    if ext == ".xlsx":
        try:
            sheets = read_xlsx_file(path)
        except Exception as e:
            all_log_lines.append(f"XLSX 읽기 실패: {e}")
            print(f"  ✗ XLSX 읽기 실패: {e}")
            return

        for sheet_idx, (sheet_name, rows, meta) in enumerate(sheets):
            all_log_lines.append(f"\n{'─' * 70}")
            all_log_lines.append(f"시트 [{sheet_idx}]: {sheet_name}")
            all_log_lines.append(f"{'─' * 70}")

            analyzer = TableAnalyzer(rows, meta)
            analyzer.analyze()
            all_log_lines.extend(analyzer.log_lines)

            cleaned = analyzer.get_cleaned_rows()
            if sheet_idx == 0:
                combined_clean_rows = cleaned
            else:
                # 멀티시트: 시트명을 표시하는 구분행 추가 후 append
                if combined_clean_rows and cleaned:
                    sep_row = [f"--- Sheet: {sheet_name} ---"] + [""] * (len(cleaned[0]) - 1) if cleaned[0] else []
                    # 열 수가 다르면 맞추기
                    target = len(combined_clean_rows[0]) if combined_clean_rows else 0
                    if cleaned:
                        for row in cleaned:
                            if len(row) < target:
                                row.extend([""] * (target - len(row)))
                            elif len(row) > target:
                                row[:] = row[:target]
                    combined_clean_rows.append([""] * target)  # 빈 행
                    combined_clean_rows.append(
                        ([f"[Sheet: {sheet_name}]"] + [""] * (target - 1)) if target > 0 else [f"[Sheet: {sheet_name}]"]
                    )
                    combined_clean_rows.extend(cleaned)
                elif cleaned:
                    combined_clean_rows = cleaned

    elif ext == ".csv":
        try:
            rows, meta = read_csv_file(path)
        except Exception as e:
            all_log_lines.append(f"CSV 읽기 실패: {e}")
            print(f"  ✗ CSV 읽기 실패: {e}")
            return

        analyzer = TableAnalyzer(rows, meta)
        analyzer.analyze()
        all_log_lines.extend(analyzer.log_lines)
        combined_clean_rows = analyzer.get_cleaned_rows()

    else:
        print(f"  ⏭ 지원하지 않는 확장자: {ext}")
        return

    # 결과 저장
    RESULT_DIR.mkdir(parents=True, exist_ok=True)

    # Cleaned CSV
    csv_path = RESULT_DIR / f"{stem}.csv"
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(combined_clean_rows)
    print(f"  ✓ Cleaned CSV: {csv_path.name} ({len(combined_clean_rows)} rows)")

    # Analysis log
    txt_path = RESULT_DIR / f"{stem}.txt"
    txt_path.write_text("\n".join(all_log_lines), encoding="utf-8")
    print(f"  ✓ Analysis log: {txt_path.name} ({len(all_log_lines)} lines)")


# ──────────────────────────────────────────
# Main
# ──────────────────────────────────────────

def main():
    print(f"\n{'═' * 60}")
    print(f"  PyNorma Specimen Analysis Pipeline")
    print(f"  Source: {SPECIMEN_DIR}")
    print(f"  Output: {RESULT_DIR}")
    print(f"{'═' * 60}")

    # 대상 파일 수집
    files = sorted(SPECIMEN_DIR.glob("*"))
    files = [f for f in files if f.is_file()
             and not f.name.startswith("_")
             and not f.name.startswith(".")
             and f.suffix.lower() in (".csv", ".xlsx")]

    print(f"\n  대상 파일: {len(files)}개")

    for fpath in files:
        try:
            process_file(fpath)
        except Exception as e:
            print(f"  ✗ {fpath.name}: {e}")
            import traceback
            traceback.print_exc()

    # Summary
    print(f"\n{'═' * 60}")
    print(f"  분석 완료!")
    print(f"{'═' * 60}")

    results = sorted(RESULT_DIR.glob("*"))
    csv_files = [f for f in results if f.suffix.lower() == ".csv"]
    txt_files = [f for f in results if f.suffix.lower() == ".txt"]
    print(f"  Cleaned CSV: {len(csv_files)}개")
    print(f"  Analysis log: {len(txt_files)}개")

    for f in sorted(results):
        print(f"    {f.name:50s} {f.stat().st_size:>10,} bytes")


if __name__ == "__main__":
    main()
