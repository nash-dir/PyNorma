"""
specimen/_collect_specimens.py
==============================
전처리 실험용 '더러운 테이블 데이터' 샘플을 수집·생성하는 스크립트.

수집 전략
---------
1. GitHub raw 파일 직접 다운로드 (공개 CSV)
2. Python으로 다양한 패턴의 더러운 데이터를 합성 생성 (CSV + XLSX)

보안(Sanitize)
--------------
- 다운로드 파일: 텍스트(CSV)만 허용, =/@/+/- 로 시작하는 셀 무력화 (CSV Injection 방지)
- XLSX 생성: openpyxl 기본 모드(매크로·VBA 없음)
- 바이너리 실행파일 다운로드 금지
"""

from __future__ import annotations

import csv
import io
import os
import random
import re
import string
import sys
import urllib.request
import urllib.error
from pathlib import Path
from datetime import datetime, timedelta

SPECIMEN_DIR = Path(__file__).resolve().parent


# ─────────────────────────────────────────────
# Sanitization Helpers
# ─────────────────────────────────────────────

def sanitize_csv_cell(value: str) -> str:
    """CSV Injection 방지: 위험 접두어 제거."""
    if isinstance(value, str):
        # Strip BOM / null bytes
        value = value.replace("\x00", "").replace("\ufeff", "")
        # Neutralize formula injection
        if value and value[0] in ("=", "+", "-", "@", "\t", "\r", "\n"):
            value = "'" + value
    return value


def sanitize_csv_content(raw: str) -> str:
    """CSV 텍스트 전체에 대해 셀 단위 sanitize."""
    reader = csv.reader(io.StringIO(raw))
    out = io.StringIO()
    writer = csv.writer(out)
    for row in reader:
        writer.writerow([sanitize_csv_cell(c) for c in row])
    return out.getvalue()


def is_safe_text(raw: bytes) -> bool:
    """바이너리/실행파일이 아닌 텍스트인지 확인."""
    # Check for common binary signatures
    binary_sigs = [b"\x50\x4b\x03\x04",  # ZIP/XLSX/DOCX
                   b"\x4d\x5a",           # PE exe
                   b"\x7fELF",            # ELF
                   b"\xd0\xcf\x11\xe0",   # OLE (old xls with macros)
                   b"\x25\x50\x44\x46"]   # PDF
    for sig in binary_sigs:
        if raw[:4].startswith(sig):
            return False
    # Check if mostly text
    try:
        raw.decode("utf-8", errors="strict")
        return True
    except UnicodeDecodeError:
        try:
            raw.decode("latin-1", errors="strict")
            text_chars = sum(1 for b in raw if 32 <= b < 127 or b in (9, 10, 13))
            return text_chars / max(len(raw), 1) > 0.85
        except Exception:
            return False


# ─────────────────────────────────────────────
# Part 1: Download CSV from GitHub
# ─────────────────────────────────────────────

DOWNLOAD_SOURCES: list[dict] = [
    {
        "url": "https://raw.githubusercontent.com/ryanleeallred/datasets/master/messy-data.csv",
        "filename": "github_messy_data.csv",
        "desc": "Messy data with mixed types, missing values",
    },
    {
        "url": "https://raw.githubusercontent.com/quankiquanki/skyline-problem/master/data/dirty_data.csv",
        "filename": "github_dirty_skyline.csv",
        "desc": "Dirty skyline problem data",
    },
    {
        "url": "https://raw.githubusercontent.com/Quansight/qhub/main/tests/_cli/data/dirty.csv",
        "filename": "github_qhub_dirty.csv",
        "desc": "QHub dirty test data",
    },
]


def download_csv_files():
    """GitHub에서 CSV 파일 다운로드 + sanitize."""
    print("=" * 60)
    print("Part 1: Downloading CSV files from GitHub")
    print("=" * 60)

    headers = {"User-Agent": "Mozilla/5.0 (PyNorma-Specimen-Collector)"}
    downloaded = 0

    for src in DOWNLOAD_SOURCES:
        dest = SPECIMEN_DIR / src["filename"]
        print(f"\n  → {src['desc']}")
        print(f"    URL: {src['url']}")

        try:
            req = urllib.request.Request(src["url"], headers=headers)
            with urllib.request.urlopen(req, timeout=15) as resp:
                raw = resp.read()

            # Safety check
            if not is_safe_text(raw):
                print(f"    ⚠ SKIPPED: Binary or unsafe content detected!")
                continue

            text = raw.decode("utf-8", errors="replace")
            sanitized = sanitize_csv_content(text)

            dest.write_text(sanitized, encoding="utf-8")
            print(f"    ✓ Saved ({len(sanitized):,} bytes) → {dest.name}")
            downloaded += 1
        except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as e:
            print(f"    ✗ Download failed: {e}")
        except Exception as e:
            print(f"    ✗ Error: {e}")

    print(f"\n  Downloaded: {downloaded}/{len(DOWNLOAD_SOURCES)}")


# ─────────────────────────────────────────────
# Part 2: Generate Synthetic Dirty Data
# ─────────────────────────────────────────────

random.seed(42)


def _rand_name():
    first = random.choice(["김", "이", "박", "최", "정", "John", "Jane", "Bob", "Alice", "田中"])
    last = random.choice(["민수", "영희", "철수", "유진", "Smith", "Doe", "Lee", "太郎", "花子", ""])
    return f"{first}{last}"


def _rand_date(fmt=None):
    """랜덤 날짜를 일관성 없는 포맷으로 생성."""
    base = datetime(2020, 1, 1) + timedelta(days=random.randint(0, 1500))
    formats = [
        "%Y-%m-%d", "%d/%m/%Y", "%m-%d-%Y", "%Y.%m.%d",
        "%d %b %Y", "%B %d, %Y", "%Y%m%d", "%d-%m-%y",
    ]
    return base.strftime(fmt or random.choice(formats))


def _rand_amount():
    """금액: 가끔 단위·통화 문자 섞임."""
    val = round(random.uniform(100, 99999), 2)
    decorations = [
        f"{val}", f"${val}", f"₩{int(val)}", f"{val}원",
        f"{val:,.2f}", f"  {val}  ", f"-{val}", f"({val})",
        str(int(val)), f"{val} USD",
    ]
    return random.choice(decorations)


def _maybe_missing(value, prob=0.15):
    """확률적으로 결측값."""
    if random.random() < prob:
        return random.choice(["", "N/A", "n/a", "-", "NULL", "null", "NaN", "?", "  ", "미입력"])
    return value


def gen_01_messy_sales_csv():
    """01: 판매 데이터 - 중복행, 결측, 불일치 날짜/금액 포맷."""
    path = SPECIMEN_DIR / "01_messy_sales.csv"
    print(f"\n  → Generating {path.name}")

    headers = ["주문번호", "고객명", "주문일", "제품", "수량", "단가", "합계", "비고"]
    products = ["노트북", "키보드", "마우스", "모니터", "USB 허브", "Laptop", "Keyboard", "Mouse"]

    rows = [headers]
    for i in range(200):
        order_id = _maybe_missing(f"ORD-{1000 + i:04d}")
        name = _maybe_missing(_rand_name())
        date = _maybe_missing(_rand_date())
        product = _maybe_missing(random.choice(products))
        qty = _maybe_missing(str(random.randint(1, 50)))
        price = _maybe_missing(_rand_amount())
        total = _maybe_missing(_rand_amount())
        note = _maybe_missing(random.choice(["정상", "반품", "교환", "할인적용", "", "  ", "취소??"]))

        row = [order_id, name, date, product, qty, price, total, note]
        rows.append(row)

        # 5% 확률로 중복행
        if random.random() < 0.05:
            rows.append(row.copy())

    # 중간에 빈 행 삽입
    for _ in range(10):
        pos = random.randint(1, len(rows) - 1)
        rows.insert(pos, [""] * len(headers))

    # 끝에 합계행 (실제 데이터에서 흔히 보이는 패턴)
    rows.append(["", "", "", "", "", "", "총합계: ₩2,345,678", ""])
    rows.append(["", "", "", "생성일: " + _rand_date("%Y-%m-%d"), "", "", "", ""])

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    print(f"    ✓ {len(rows)} rows")


def gen_02_multiheader_csv():
    """02: 다중 헤더 + 부제목 행이 섞인 CSV."""
    path = SPECIMEN_DIR / "02_multiheader_report.csv"
    print(f"\n  → Generating {path.name}")

    rows = []
    rows.append(["", "", "분기별 매출 보고서", "", "", ""])
    rows.append(["", "", "2024년 1~4분기", "", "", ""])
    rows.append([""])
    rows.append(["지역", "담당자", "Q1", "Q2", "Q3", "Q4"])
    rows.append(["", "", "(만원)", "(만원)", "(만원)", "(만원)"])

    regions = ["서울", "부산", "대구", "인천", "광주", "대전", "Seoul", "Busan"]
    for region in regions:
        rows.append([
            region,
            _rand_name(),
            _maybe_missing(str(random.randint(100, 9999))),
            _maybe_missing(str(random.randint(100, 9999))),
            _maybe_missing(str(random.randint(100, 9999))),
            _maybe_missing(str(random.randint(100, 9999))),
        ])

    rows.append([""])
    rows.append(["소계", "", "32,150", "28,400", "35,200", "41,000"])
    rows.append(["", "", "", "", "", ""])
    rows.append(["※ 본 자료는 내부 참고용입니다.", "", "", "", "", ""])

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    print(f"    ✓ {len(rows)} rows")


def gen_03_mixed_encoding_csv():
    """03: 인코딩 혼란 - BOM, 비ASCII 문자, 이스케이프 문제."""
    path = SPECIMEN_DIR / "03_encoding_chaos.csv"
    print(f"\n  → Generating {path.name}")

    rows = []
    rows.append(["ID", "Name", "City", "Comment"])

    entries = [
        ["1", "José García", "São Paulo", "Très bien – merci!"],
        ["2", "田中太郎", "東京", "日本語テスト　全角スペース"],
        ["3", "Müller, Hans", "München", "Ünited Königdom"],
        ["4", "Фёдоров", "Москва", "Привет мир"],
        ["5", "이영희", "서울", "한글 데이터 — 대시 종류 다양"],
        ["6", "O'Brien", "New\nYork", "Line\r\nbreak in cell"],
        ["7", '"Quoted"', "It's a \"test\"", "comma, in, cell"],
        ["8", "Tab\there", "null\x00byte", "special chars: <>&"],
        ["9", "", "  spaces  ", " leading/trailing "],
        ["10", "ALLCAPS", "lowercase", "MiXeD CaSe"],
    ]

    for e in entries:
        rows.append(e)

    # Add some with formula injection attempts (will be sanitized on read)
    rows.append(["11", "=CMD('calc')", "+1+1", "@SUM(A1:A10)"])
    rows.append(["12", "Normal", "Data", "-1+1"])

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    print(f"    ✓ {len(rows)} rows")


def gen_04_ragged_csv():
    """04: 들쭉날쭉한 열 수 (ragged rows)."""
    path = SPECIMEN_DIR / "04_ragged_columns.csv"
    print(f"\n  → Generating {path.name}")

    lines = []
    lines.append("A,B,C,D,E")
    for i in range(80):
        ncols = random.choice([3, 4, 5, 6, 7, 2, 5, 5, 5])
        vals = [str(random.randint(1, 999)) for _ in range(ncols)]
        if random.random() < 0.1:
            vals[random.randint(0, len(vals) - 1)] = ""
        lines.append(",".join(vals))

    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"    ✓ {len(lines)} rows")


def gen_05_dirty_xlsx():
    """05: XLSX - 병합 셀, 빈 행/열, 서식 불일치, 다중 시트."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("\n  ⚠ openpyxl not installed, skipping XLSX generation")
        print("    Install: pip install openpyxl")
        return

    path = SPECIMEN_DIR / "05_merged_cells_mess.xlsx"
    print(f"\n  → Generating {path.name}")

    wb = openpyxl.Workbook()

    # Sheet 1: 병합 셀 + 소계행
    ws1 = wb.active
    ws1.title = "매출_원본"

    # 타이틀 (병합)
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "2024년 상반기 매출 보고"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.merge_cells("A2:F2")
    ws1["A2"] = "작성일: 2024-07-15"

    # 빈 행
    # 헤더 (4행)
    headers = ["지역", "지점", "1월", "2월", "3월", "비고"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", fill_type="solid")

    row_num = 5
    regions = {"서울": ["강남점", "종로점", "영등포점"], "부산": ["해운대점", "서면점"], "대구": ["동성로점"]}
    for region, branches in regions.items():
        start_row = row_num
        for branch in branches:
            ws1.cell(row=row_num, column=1, value=region if branch == branches[0] else "")
            ws1.cell(row=row_num, column=2, value=branch)
            for col in range(3, 6):
                val = random.randint(500, 9999)
                ws1.cell(row=row_num, column=col, value=_maybe_missing(str(val), 0.2))
            ws1.cell(row=row_num, column=6, value=_maybe_missing(""))
            row_num += 1

        # 지역 셀 병합
        if len(branches) > 1:
            ws1.merge_cells(start_row=start_row, start_column=1,
                            end_row=start_row + len(branches) - 1, end_column=1)

        # 소계행
        ws1.cell(row=row_num, column=2, value="소계")
        ws1.cell(row=row_num, column=2).font = Font(bold=True)
        for col in range(3, 6):
            ws1.cell(row=row_num, column=col, value=random.randint(5000, 30000))
        row_num += 1
        # 빈 행
        row_num += 1

    # 하단 주석
    ws1.cell(row=row_num, column=1, value="※ 단위: 만원")
    ws1.cell(row=row_num + 1, column=1, value="※ 일부 지점 데이터 미입력")

    # Sheet 2: 타입 혼란 시트
    ws2 = wb.create_sheet("혼합_타입")
    ws2.append(["ID", "값", "날짜", "메모"])
    type_chaos = [
        [1, 100, "2024-01-01", "정상"],
        [2, "백오십", "1월 15일", "숫자가 텍스트"],
        ["3", 200.5, "2024/02/01", "ID가 텍스트"],
        [4, "N/A", "", "결측"],
        [5, True, "20240301", "불리언 혼입"],
        [6, "  300  ", " 2024-04-01 ", "공백 포함"],
        [7, -999, "9999-99-99", "이상값"],
        [None, None, None, "전부 None"],
        [8, "=1+1", "today()", "수식 텍스트"],
        [9, 0, "0", "영 vs 문자열 영"],
    ]
    for row in type_chaos:
        ws2.append(row)

    # Sheet 3: 빈 열이 중간에 있는 시트
    ws3 = wb.create_sheet("빈_열_포함")
    ws3.append(["A", "", "B", "", "C"])  # 빈 열
    for i in range(20):
        ws3.append([
            random.randint(1, 100),
            "",  # 빈 열
            _rand_name(),
            "",  # 빈 열
            _rand_amount(),
        ])

    wb.save(path)
    print(f"    ✓ 3 sheets saved")


def gen_06_pivot_like_csv():
    """06: 피벗테이블 스타일 와이드 포맷 + 소계·총계."""
    path = SPECIMEN_DIR / "06_pivot_style_table.csv"
    print(f"\n  → Generating {path.name}")

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    categories = ["Electronics", "Clothing", "Food", "Books"]

    rows = []
    rows.append(["Category / Month"] + months + ["Total"])

    for cat in categories:
        vals = [random.randint(100, 999) for _ in months]
        rows.append([cat] + [str(v) for v in vals] + [str(sum(vals))])

    # 소계
    rows.append(["Subtotal"] + [str(random.randint(500, 3000)) for _ in months] + [""])
    rows.append([""] * (len(months) + 2))
    rows.append(["Tax (10%)"] + [str(random.randint(50, 300)) for _ in months] + [""])
    rows.append(["Grand Total"] + [str(random.randint(600, 3500)) for _ in months] + [""])
    rows.append([""] * (len(months) + 2))
    rows.append(["Report generated: " + _rand_date("%Y-%m-%d %H:%M")] + [""] * (len(months) + 1))

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    print(f"    ✓ {len(rows)} rows")


def gen_07_semicolon_csv():
    """07: 세미콜론 구분자 + 유럽식 소수점(쉼표)."""
    path = SPECIMEN_DIR / "07_semicolon_european.csv"
    print(f"\n  → Generating {path.name}")

    lines = []
    lines.append("Produkt;Preis;Menge;Datum")
    products_de = ["Kartoffel", "Brot", "Milch", "Käse", "Wurst", "Apfel", "Butter"]
    for i in range(60):
        product = random.choice(products_de)
        price = f"{random.uniform(0.5, 25.0):.2f}".replace(".", ",")  # 유럽식
        qty = str(random.randint(1, 100))
        date = _rand_date("%d.%m.%Y")
        line = f"{product};{_maybe_missing(price)};{_maybe_missing(qty)};{_maybe_missing(date)}"
        lines.append(line)

    path.write_text("\n".join(lines), encoding="utf-8")
    print(f"    ✓ {len(lines)} rows")


def gen_08_annotation_heavy_xlsx():
    """08: XLSX - 사이드 주석·하단 각주가 많은 통계표."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Border, Side, Alignment
    except ImportError:
        print("\n  ⚠ openpyxl not installed, skipping XLSX generation")
        return

    path = SPECIMEN_DIR / "08_annotated_stats_table.xlsx"
    print(f"\n  → Generating {path.name}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "통계표"

    # 상단 메타데이터
    ws.merge_cells("A1:H1")
    ws["A1"] = "표 3-2. 연도별 주요 경제지표"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"] = "(단위: %, 십억원)"
    ws["A2"].font = Font(italic=True, size=9)

    # 데이터
    headers = ["연도", "GDP성장률", "물가상승률", "실업률", "수출", "수입", "경상수지", "주1)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)

    for year in range(2015, 2025):
        row = year - 2015 + 5
        ws.cell(row=row, column=1, value=year)
        ws.cell(row=row, column=2, value=_maybe_missing(f"{random.uniform(0.5, 4.5):.1f}"))
        ws.cell(row=row, column=3, value=_maybe_missing(f"{random.uniform(0.3, 5.0):.1f}"))
        ws.cell(row=row, column=4, value=_maybe_missing(f"{random.uniform(2.5, 5.5):.1f}"))
        ws.cell(row=row, column=5, value=_maybe_missing(str(random.randint(400, 700))))
        ws.cell(row=row, column=6, value=_maybe_missing(str(random.randint(350, 650))))
        ws.cell(row=row, column=7, value=_maybe_missing(str(random.randint(-50, 100))))
        # 사이드 주석 (흔한 더러운 패턴)
        if year in (2020, 2021):
            ws.cell(row=row, column=8, value="코로나 영향")
        elif year == 2023:
            ws.cell(row=row, column=8, value="잠정치")
        elif year == 2024:
            ws.cell(row=row, column=8, value="추정치p)")

    # 하단 각주
    foot_row = 16
    ws.cell(row=foot_row, column=1, value="주1) p)는 잠정치를 의미함").font = Font(size=8, italic=True)
    ws.cell(row=foot_row + 1, column=1, value="자료: 한국은행 경제통계시스템").font = Font(size=8)
    ws.cell(row=foot_row + 2, column=1, value="작성기관: 기획재정부").font = Font(size=8)

    wb.save(path)
    print(f"    ✓ Saved")


def gen_09_wide_sparse_csv():
    """09: 극단적으로 넓고 듬성듬성한 테이블."""
    path = SPECIMEN_DIR / "09_wide_sparse.csv"
    print(f"\n  → Generating {path.name}")

    ncols = 50
    headers = [f"Feature_{i:03d}" for i in range(ncols)]
    rows = [headers]
    for _ in range(40):
        row = []
        for _ in range(ncols):
            if random.random() < 0.7:  # 70% 결측
                row.append("")
            else:
                row.append(str(round(random.gauss(0, 10), 3)))
        rows.append(row)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    print(f"    ✓ {len(rows)} rows × {ncols} cols")


def gen_10_mixed_tables_xlsx():
    """10: XLSX - 하나의 시트에 여러 작은 테이블이 산재."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("\n  ⚠ openpyxl not installed, skipping XLSX generation")
        return

    path = SPECIMEN_DIR / "10_multiple_tables_one_sheet.xlsx"
    print(f"\n  → Generating {path.name}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "혼합"

    # Table 1 at A1
    ws["A1"] = "표1. 인원 현황"
    ws["A1"].font = Font(bold=True)
    ws.append(["부서", "인원", "평균연봉"])
    for dept in ["개발", "기획", "영업", "인사", "재무"]:
        ws.append([dept, random.randint(5, 50), f"{random.randint(3000, 8000)}만원"])

    # Gap
    ws.append([])
    ws.append([])

    # Table 2 at A10ish
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="표2. 월별 매출").font = Font(bold=True)
    row += 1
    for col, h in enumerate(["월", "실적", "목표", "달성률"], 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
    row += 1
    for month in range(1, 13):
        actual = random.randint(100, 500)
        target = random.randint(100, 500)
        rate = f"{actual / target * 100:.1f}%"
        for col, val in enumerate([f"{month}월", actual, target, rate], 1):
            ws.cell(row=row, column=col, value=val)
        row += 1

    # Table 3 offset to the right (E1)
    ws["E1"] = "표3. 장비 목록"
    ws["E1"].font = Font(bold=True)
    ws["E2"] = "장비명"
    ws["F2"] = "수량"
    ws["G2"] = "상태"
    items = [("노트북", 45, "양호"), ("모니터", 60, "일부 불량"), ("프린터", 8, "N/A"),
             ("서버", 3, "점검중"), ("", "", "")]
    for i, (name, qty, status) in enumerate(items, 3):
        ws[f"E{i}"] = name
        ws[f"F{i}"] = qty
        ws[f"G{i}"] = status

    wb.save(path)
    print(f"    ✓ 3 tables on one sheet")


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    print(f"\n{'='*60}")
    print(f"  PyNorma Specimen Collector")
    print(f"  Target: {SPECIMEN_DIR}")
    print(f"{'='*60}")

    SPECIMEN_DIR.mkdir(parents=True, exist_ok=True)

    # Part 1: Downloads
    download_csv_files()

    # Part 2: Synthetic dirty data
    print(f"\n{'='*60}")
    print(f"Part 2: Generating synthetic dirty data files")
    print(f"{'='*60}")

    gen_01_messy_sales_csv()
    gen_02_multiheader_csv()
    gen_03_mixed_encoding_csv()
    gen_04_ragged_csv()
    gen_05_dirty_xlsx()
    gen_06_pivot_like_csv()
    gen_07_semicolon_csv()
    gen_08_annotation_heavy_xlsx()
    gen_09_wide_sparse_csv()
    gen_10_mixed_tables_xlsx()

    # Summary
    print(f"\n{'='*60}")
    print(f"  Collection complete!")
    print(f"{'='*60}")
    files = sorted(SPECIMEN_DIR.glob("*"))
    files = [f for f in files if f.name != "_collect_specimens.py" and not f.name.startswith(".")]
    print(f"\n  Files in {SPECIMEN_DIR}:")
    for f in files:
        size = f.stat().st_size
        print(f"    {f.name:45s} {size:>10,} bytes")
    print(f"\n  Total: {len(files)} specimen files")


if __name__ == "__main__":
    main()
