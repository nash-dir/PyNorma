"""XLSX file parser for PyNorma."""

import logging
from typing import Dict, Literal, Optional, Tuple, Union

import openpyxl
import openpyxl.worksheet.worksheet
import pandas as pd

from pynorma.io.trimmer import trim_dataframe
from pynorma.utils import clean_dataframe
from pynorma.detect.header_finder import detect_header_end_row

logger = logging.getLogger("pynorma")


def _unmerge_cells(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    """Unmerge every merged-cell range and fill each cell with the top-left value."""
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value


def parse_xlsx(
    filepath: str,
    trim: Union[bool, Literal["auto"], dict] = "auto",
    is_header: Union[bool, int] = True,
    sheet_name: Union[str, int, None] = None,
    encoding: str = "auto",
    verbose: bool = False,
) -> Tuple[pd.DataFrame, Dict]:
    """Parse an XLSX file and return the preprocessed DataFrame with metadata.

    The function signature mirrors :func:`csv_parser.parse_csv` so that
    the top-level ``parse()`` can call both sub-parsers uniformly.

    Parameters
    ----------
    filepath : str
        Path to the Excel file.
    trim : bool or ``"auto"`` or dict
        Trimming mode passed to the trimmer.
    is_header : bool or int
        Header detection mode.
    sheet_name : str or int or None
        Sheet to read.  ``None`` selects the active (first) sheet.
    encoding : str
        Reserved for interface consistency (openpyxl handles encoding
        internally).
    verbose : bool
        Reserved for interface consistency with the top-level parser.

    Returns
    -------
    tuple of (pd.DataFrame, dict)
        The preprocessed DataFrame and a metadata dict for verbose output.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)

        if sheet_name is None:
            ws = wb.active
        elif isinstance(sheet_name, str):
            ws = wb[sheet_name]
        elif isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            raise TypeError("'sheet_name' must be a string, an integer, or None.")
    except Exception as e:
        logger.error("Failed to open Excel file '%s': %s", filepath, e)
        raise

    detected_sheet_name = ws.title

    # 1. Handle Excel-specific merged cells
    _unmerge_cells(ws)

    # 2. Build raw DataFrame from cell values
    data = ws.values
    df_raw = pd.DataFrame(data)

    # 3. Basic cleaning
    df_raw = clean_dataframe(df_raw)

    # 4. Determine header position
    header_index: Optional[int] = None
    if isinstance(is_header, int):
        header_index = is_header
    elif is_header is True:
        header_index = detect_header_end_row(df_raw)
        if header_index == -1:
            header_index = None

    # 5. Delegate to the trimmer
    df_clean, info = trim_dataframe(
        df=df_raw,
        trim_mode=trim,
        set_header=(is_header is not False),
        header_row=header_index,
    )

    info["detected_sheet"] = detected_sheet_name
    return df_clean, info