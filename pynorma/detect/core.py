"""
Core v2: TableRegion-based detection framework.

전처리를 두 단계로 분리:
  Phase 1 (detect): grid → list[TableRegion]  ← 전략이 경쟁
  Phase 2 (clean):  region → cleaned rows      ← 공통 파이프라인
"""
from __future__ import annotations

import csv
import io
import math
import re
from abc import ABC, abstractmethod
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import NamedTuple


# ═══════════════════════════════════════════
# TableRegion: the core abstraction
# ═══════════════════════════════════════════

class TableRegion(NamedTuple):
    """N × 5 integers describing one table in a grid.

    header: row index of column headers (-1 if no header)
    top:    first data row (inclusive)
    left:   first data column (inclusive)
    bottom: last data row (inclusive)
    right:  last data column (exclusive)
    """
    header: int
    top: int
    left: int
    bottom: int
    right: int


# ═══════════════════════════════════════════
# Strategy ABC
# ═══════════════════════════════════════════

class Strategy(ABC):
    name: str = "unnamed"
    description: str = ""

    @abstractmethod
    def detect(self, grid: list[list[str]]) -> list[TableRegion]:
        """Detect table regions in a raw 2D grid.

        Returns a list of TableRegion, one per detected table.
        """
        ...


# ═══════════════════════════════════════════
# Cell utilities
# ═══════════════════════════════════════════

MISSING_MARKERS = frozenset({
    "", "N/A", "n/a", "NA", "na", "NULL", "null", "NaN", "nan",
    "None", "none", "-", "--", "?", "미입력", "#N/A", "#REF!", ".", "..",
})

SUMMARY_KEYWORDS = frozenset({
    "합계", "총합계", "소계", "subtotal", "total", "grand total",
    "sum", "average", "평균", "※", "주)", "자료:", "작성", "단위:",
    "생성일", "report generated", "tax",
})


def is_empty(val: str) -> bool:
    v = str(val).strip()
    return v == "" or v in MISSING_MARKERS


def is_numeric(val: str) -> bool:
    v = str(val).strip()
    v = re.sub(r"^['\"\$₩]|[원만USD%\s]+$", "", v)
    v = v.strip("() ")
    if not v:
        return False
    try:
        float(v.replace(",", ""))
        return True
    except ValueError:
        return False


def is_date_like(val: str) -> bool:
    pats = [r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}",
            r"\d{1,2}[-/.]\d{1,2}[-/.]\d{2,4}", r"\d{8}"]
    return any(re.search(p, str(val)) for p in pats)


def cell_type(val: str) -> str:
    v = str(val).strip()
    if is_empty(v):
        return "empty"
    if is_numeric(v):
        return "numeric"
    if is_date_like(v):
        return "date"
    return "text"


def row_fill_rate(row: list[str]) -> float:
    if not row:
        return 0.0
    return sum(1 for c in row if not is_empty(c)) / len(row)


def row_fill_count(row: list[str]) -> int:
    return sum(1 for c in row if not is_empty(c))


def clean_cell(val: str) -> str:
    v = str(val).strip().replace("\x00", "").replace("\ufeff", "")
    if v.startswith("'") and len(v) > 1 and v[1] in "=+@-":
        v = v[1:]
    if v in MISSING_MARKERS:
        return ""
    return v


def shannon_entropy(counts: dict) -> float:
    total = sum(counts.values())
    if total == 0:
        return 0.0
    return -sum((c / total) * math.log2(c / total) for c in counts.values() if c > 0)


def grid_cols(grid: list[list[str]]) -> int:
    return max((len(r) for r in grid), default=0)


def grid_cell(grid: list[list[str]], r: int, c: int) -> str:
    if r < len(grid) and c < len(grid[r]):
        return grid[r][c]
    return ""


def has_summary_keyword(row: list[str]) -> bool:
    text = " ".join(str(c).strip().lower() for c in row)
    return any(kw in text for kw in SUMMARY_KEYWORDS)


def split_tables_by_gap(
    grid: list[list[str]],
    *,
    min_gap: int = 2,
    min_table_rows: int = 2,
) -> list[TableRegion]:
    """Detect multiple tables by finding empty-row gaps.

    Scans for consecutive empty rows (fill_rate == 0) of length ≥ min_gap.
    Each dense segment is treated as a separate table.

    Parameters
    ----------
    grid : list of list of str
        Raw 2D grid.
    min_gap : int
        Minimum number of consecutive empty rows to count as a table separator.
    min_table_rows : int
        Minimum data rows for a segment to be considered a table.

    Returns
    -------
    list of TableRegion
        One region per detected table.
    """
    if not grid:
        return []

    n = len(grid)
    ncols = grid_cols(grid)

    # Classify each row as empty or not
    fill = [row_fill_rate(grid[i]) > 0 for i in range(n)]

    # Find segments of non-empty rows separated by gaps
    segments = []
    seg_start = None
    empty_count = 0

    for i in range(n):
        if fill[i]:
            if seg_start is None:
                seg_start = i
            empty_count = 0
        else:
            empty_count += 1
            if seg_start is not None and empty_count >= min_gap:
                seg_end = i - empty_count + 1
                if seg_end > seg_start:
                    segments.append((seg_start, seg_end - 1))
                seg_start = None
                empty_count = 0

    # Final segment
    if seg_start is not None:
        seg_end = n - 1
        while seg_end > seg_start and not fill[seg_end]:
            seg_end -= 1
        if seg_end >= seg_start:
            segments.append((seg_start, seg_end))

    if not segments:
        return []

    # If only 1 segment, return single table (no split needed)
    if len(segments) == 1:
        return []  # Let per-strategy detection handle it

    # Convert segments to TableRegions
    regions = []
    for seg_start, seg_end in segments:
        data_rows = seg_end - seg_start + 1
        if data_rows < min_table_rows:
            continue

        # Header: first text-dominant row in segment
        header = seg_start
        for i in range(seg_start, min(seg_start + 10, seg_end + 1)):
            non_empty = [c for c in grid[i] if not is_empty(c)]
            if len(non_empty) < 2:
                continue
            text_count = sum(1 for c in non_empty if not is_numeric(c))
            if text_count / len(non_empty) > 0.5:
                header = i
                break

        top = header + 1
        if top > seg_end:
            top = seg_end

        # Column boundaries
        data_idxs = [i for i in range(top, seg_end + 1) if row_fill_rate(grid[i]) > 0]
        left, right = 0, ncols
        if data_idxs:
            for c in range(ncols):
                filled = sum(1 for i in data_idxs if c < len(grid[i]) and not is_empty(grid[i][c]))
                if filled / len(data_idxs) >= 0.1:
                    left = c
                    break
            for c in range(ncols - 1, left - 1, -1):
                filled = sum(1 for i in data_idxs if c < len(grid[i]) and not is_empty(grid[i][c]))
                if filled / len(data_idxs) >= 0.1:
                    right = c + 1
                    break

        regions.append(TableRegion(header=header, top=top, left=left,
                                   bottom=seg_end, right=right))

    return regions


# ═══════════════════════════════════════════
# Recursive layout segmentation (XY-cut)
# ═══════════════════════════════════════════
#
# A sheet may hold several tables laid out side-by-side (separated by an empty
# column band) and/or stacked (separated by a blank-row band OR a section-title
# / repeated-header row). We recursively cut the grid into content blocks, then
# detect one table per block. Cuts are gated so a *single* table with interior
# empty columns/rows (e.g. 18_empty_cols_middle) is NOT split.

def _cells_content(grid, r0, r1, c0, c1) -> "tuple[set, set]":
    """Return (rows_with_content, cols_with_content) within a window."""
    rows, cols = set(), set()
    for r in range(r0, r1 + 1):
        row = grid[r] if r < len(grid) else []
        for c in range(c0, c1 + 1):
            if c < len(row) and not is_empty(row[c]):
                rows.add(r)
                cols.add(c)
    return rows, cols


def _side_titled(grid, r0, r1, ca, cb) -> bool:
    """True if this column band has its own caption near its top — a title row
    confined to columns [ca, cb]. Two side-by-side tables each carry their own
    caption (표1 … 표3); one table with a spacer column does not."""
    for r in range(r0, min(r0 + 3, r1 + 1)):
        if _is_title_row(grid, r, ca, cb):
            return True
    return False


def _find_col_gap(grid, r0, r1, c0, c1):
    """Find a fully-empty column band that separates two *different* tables.

    Returns ((la, lb), (ra, rb)) column sub-ranges, or None. Guards against
    splitting one table with interior blank columns (18_empty_cols_middle):
    the split is accepted only when the two sides look like independent tables —
    either they occupy substantially different row sets (low row-overlap, as
    with tables of different heights) OR each side carries its own caption row.
    """
    empty_col = []
    for c in range(c0, c1 + 1):
        has = any(r < len(grid) and c < len(grid[r]) and not is_empty(grid[r][c])
                  for r in range(r0, r1 + 1))
        empty_col.append(not has)

    c = 0
    n = c1 - c0 + 1
    while c < n and empty_col[c]:
        c += 1
    while c < n:
        while c < n and not empty_col[c]:
            c += 1
        if c >= n:
            break  # no gap after this run
        gap_start = c
        while c < n and empty_col[c]:
            c += 1
        if c >= n:
            break  # trailing empties, not a separating gap
        left_a, left_b = c0, c0 + gap_start - 1
        right_a, right_b = c0 + c, c1
        lrows, _ = _cells_content(grid, r0, r1, left_a, left_b)
        rrows, _ = _cells_content(grid, r0, r1, right_a, right_b)
        if lrows and rrows:
            overlap = len(lrows & rrows) / len(lrows | rrows)
            titled = (_side_titled(grid, r0, r1, left_a, left_b)
                      and _side_titled(grid, r0, r1, right_a, right_b))
            if overlap < 0.7 or titled:
                return (left_a, left_b), (right_a, right_b)
    return None


def _row_signature(grid, r, c0, c1) -> tuple:
    """Coarse type signature of a row's value cells — for repeated-header ID."""
    sig = []
    for c in range(c0, c1 + 1):
        v = clean_cell(grid_cell(grid, r, c))
        sig.append(v.lower() if v else "")
    return tuple(sig)


def _is_title_row(grid, r, c0, c1) -> bool:
    """A lone section caption: text in the first column(s), value zone empty."""
    vals = [grid_cell(grid, r, c) for c in range(c0, c1 + 1)]
    non_empty = [(i, v) for i, v in enumerate(vals) if not is_empty(v)]
    if not (1 <= len(non_empty) <= 2):
        return False
    # all non-empty cells are in the first two columns and are text
    if any(i > 1 for i, _ in non_empty):
        return False
    return all(not is_numeric(v) for _, v in non_empty)


def _find_row_boundary(grid, r0, r1, c0, c1):
    """Find a row where a NEW stacked table begins, or None.

    Two boundary kinds, both requiring real data above the boundary:
      * blank-row band (≥1 empty row followed by content)
      * section start: a title row (or repeated header) that begins a fresh
        header→data block, with a distinct data table already present above.
    """
    content_rows, _ = _cells_content(grid, r0, r1, c0, c1)
    if not content_rows:
        return None
    first = min(content_rows)

    def is_blank(r: int) -> bool:
        return all(is_empty(grid_cell(grid, r, c)) for c in range(c0, c1 + 1))

    def header_like(r: int) -> bool:
        return (_row_datalike_frac(grid, r, c0, c1) < 0.5
                and sum(1 for c in range(c0, c1 + 1)
                        if not is_empty(grid_cell(grid, r, c))) >= 2)

    # Header signature of the first table (to detect exact repeats below).
    first_header_sig = None
    for r in range(first, min(first + 6, r1 + 1)):
        if header_like(r):
            first_header_sig = _row_signature(grid, r, c0, c1)
            break

    seen_data = False
    r = first + 1
    while r <= r1:
        row_content = not is_blank(r)
        if row_content:
            seen_data = True

        # Blank-row band (≥2) after real data → the next content row starts a
        # new stacked table.
        if seen_data and is_blank(r):
            gap = r
            while gap <= r1 and is_blank(gap):
                gap += 1
            if gap - r >= 2 and gap <= r1:
                return gap
            r = gap
            continue

        if seen_data and row_content:
            # Repeated header row → new table starts here.
            if (first_header_sig is not None
                    and _row_signature(grid, r, c0, c1) == first_header_sig):
                return r
            # Section title row introducing a fresh header+data block.
            if _is_title_row(grid, r, c0, c1):
                if any(header_like(rr) for rr in range(r + 1, min(r + 4, r1 + 1))):
                    return r
        r += 1
    return None


def segment_blocks(grid, r0=None, r1=None, c0=None, c1=None, *, _depth=0):
    """Recursively cut a grid window into content blocks (row/col ranges).

    Returns a list of ``(r0, r1, c0, c1)`` tuples, one per detected table
    block. Column cuts (side-by-side) take priority over row cuts (stacked).
    """
    if r0 is None:
        r0, c0 = 0, 0
        r1 = len(grid) - 1
        c1 = grid_cols(grid) - 1
    if r0 > r1 or c0 > c1 or _depth > 40:
        return []

    rows, cols = _cells_content(grid, r0, r1, c0, c1)
    if not rows or not cols:
        return []
    r0, r1 = min(rows), max(rows)
    c0, c1 = min(cols), max(cols)

    col_gap = _find_col_gap(grid, r0, r1, c0, c1)
    if col_gap is not None:
        (la, lb), (ra, rb) = col_gap
        return (segment_blocks(grid, r0, r1, la, lb, _depth=_depth + 1)
                + segment_blocks(grid, r0, r1, ra, rb, _depth=_depth + 1))

    row_bnd = _find_row_boundary(grid, r0, r1, c0, c1)
    if row_bnd is not None:
        return (segment_blocks(grid, r0, row_bnd - 1, c0, c1, _depth=_depth + 1)
                + segment_blocks(grid, row_bnd, r1, c0, c1, _depth=_depth + 1))

    return [(r0, r1, c0, c1)]


# ═══════════════════════════════════════════
# File reader
# ═══════════════════════════════════════════

# Non-standard delimiters, tried only when no standard delimiter is present and
# the split is consistent across lines (so single-column text isn't mis-split).
_NONSTD_DELIMS = ["::", r"\s{2,}"]


def _consistent_split(lines: list[str], pattern: str) -> bool:
    """True if `pattern` splits most lines into the same number (>1) of fields."""
    counts = [len(re.split(pattern, ln.strip())) for ln in lines if ln.strip()]
    if len(counts) < 3:
        return False
    modal, freq = Counter(counts).most_common(1)[0]
    return modal >= 2 and freq / len(counts) >= 0.8


def guess_delimiter(text: str) -> str:
    all_lines = text.strip().split("\n")
    scores = {",": 0, ";": 0, "\t": 0, "|": 0}
    for line in all_lines[:5]:
        for sep in scores:
            scores[sep] += line.count(sep)
    best = max(scores, key=scores.get)
    if scores[best] > 0:
        return best
    # No standard delimiter — try non-standard, but only if it splits most
    # lines into a consistent (>1) field count (else it is single-column text).
    sample = [ln for ln in all_lines if ln.strip()][:20]
    for cand in _NONSTD_DELIMS:
        if _consistent_split(sample, cand):
            return cand
    return ","


def _select_sheet(wb, sheet=None):
    """Pick which worksheet to read.

    Honors an explicit `sheet` (name or 0-based index). Otherwise keeps the
    first sheet, overriding only when it is a near-empty cover/readme sheet and
    another sheet clearly holds more tabular content.
    """
    names = wb.sheetnames
    if not names:
        return None
    if sheet is not None:
        if isinstance(sheet, bool):
            pass
        elif isinstance(sheet, int) and 0 <= sheet < len(names):
            return names[sheet]
        elif sheet in names:
            return sheet

    def table_score(name: str, cap: int = 8) -> int:
        """Count rows with >= 2 non-empty cells — a proxy for tabular content
        (a cover/readme sheet is mostly single-cell note rows)."""
        score = 0
        for row in wb[name].iter_rows(values_only=True):
            if sum(1 for c in row if c is not None and str(c).strip() != "") >= 2:
                score += 1
                if score >= cap:
                    break
        return score

    first = names[0]
    if len(names) == 1 or table_score(first) >= 2:
        return first  # first sheet is already tabular (or the only sheet) → keep it
    return max(names, key=table_score)


def read_specimen(path: Path, sheet=None) -> tuple[list[list[str]], dict]:
    """Read CSV/XLSX into 2D string grid + meta.

    For multi-sheet XLSX, `sheet` (name or 0-based index) forces a worksheet;
    otherwise the first sheet is used unless it is a near-empty cover sheet.
    """
    ext = path.suffix.lower()
    meta = {"filename": path.name, "format": ext, "size": path.stat().st_size}

    if ext == ".csv":
        raw = path.read_bytes()
        for enc in ["utf-8-sig", "utf-8", "cp949", "latin-1"]:
            try:
                text = raw.decode(enc)
                meta["encoding"] = enc
                break
            except (UnicodeDecodeError, LookupError):
                continue
        else:
            text = raw.decode("latin-1", errors="replace")
            meta["encoding"] = "latin-1"
        delim = guess_delimiter(text)
        meta["delimiter"] = delim
        if delim in (",", ";", "\t", "|"):
            rows = list(csv.reader(io.StringIO(text), delimiter=delim))
        else:
            # non-standard (regex) delimiter — split each line individually
            rows = [re.split(delim, ln.strip()) if ln.strip() else []
                    for ln in text.splitlines()]
    elif ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        name = _select_sheet(wb, sheet)
        ws = wb[name]
        rows = [[str(c) if c is not None else "" for c in row]
                for row in ws.iter_rows(values_only=True)]
        wb.close()
        meta["sheet"] = name if name else ""
    else:
        rows = []

    meta["rows"] = len(rows)
    meta["cols"] = grid_cols(rows)
    return rows, meta


# ═══════════════════════════════════════════
# Phase 2: Common cleaner
# ═══════════════════════════════════════════

def clean_region(grid: list[list[str]], region: TableRegion) -> list[list[str]]:
    """Extract and clean a TableRegion from the grid.

    Returns a 2D list: [header_row, data_row_1, data_row_2, ...].
    Applies: cell cleaning, empty-row removal, duplicate removal, summary-row removal.
    """
    h, top, left, bottom, right = region

    # Header
    if h >= 0 and h < len(grid):
        headers = [clean_cell(grid_cell(grid, h, c)) or f"Col_{c - left}"
                   for c in range(left, right)]
    else:
        headers = [f"Col_{i}" for i in range(right - left)]

    result = [headers]
    seen = set()

    for r in range(top, min(bottom + 1, len(grid))):
        row = [clean_cell(grid_cell(grid, r, c)) for c in range(left, right)]

        # Skip empty rows
        if all(v == "" for v in row):
            continue

        # Skip summary/annotation rows
        raw_row = [grid_cell(grid, r, c) for c in range(left, right)]
        if has_summary_keyword(raw_row):
            continue

        # Skip duplicates
        key = tuple(row)
        if key in seen:
            continue
        seen.add(key)

        result.append(row)

    return result


# ═══════════════════════════════════════════
# Structure model: TableRegion → TableModel
# ═══════════════════════════════════════════

class TableModel(NamedTuple):
    """Full structural model of one table — everything long-form conversion needs.

    header_rows: (first, last) inclusive row indices of the header block,
                 or None for headerless tables.
    stub_end:    exclusive column index where the stub (row-label columns) end.
                 stub_end == left means no stub columns.
    top/left/bottom/right: data block. top is the first row AFTER the header
                 block; right is exclusive (same convention as TableRegion).
    """
    header_rows: "tuple[int, int] | None"
    stub_end: int
    top: int
    left: int
    bottom: int
    right: int


def _row_datalike_frac(grid: list[list[str]], r: int, left: int, right: int) -> float:
    """Fraction of non-empty cells in row r (cols left..right) that are numeric/date."""
    vals = [grid_cell(grid, r, c) for c in range(left, right)]
    non_empty = [v for v in vals if not is_empty(v)]
    if not non_empty:
        return 0.0
    datalike = sum(1 for v in non_empty if is_numeric(v) or is_date_like(v))
    return datalike / len(non_empty)


def _col_datalike_frac(grid: list[list[str]], top: int, bottom: int, c: int) -> float:
    """Fraction of non-empty cells in column c (rows top..bottom) that are numeric/date."""
    vals = [grid_cell(grid, r, c) for r in range(top, bottom + 1)]
    non_empty = [v for v in vals if not is_empty(v)]
    if not non_empty:
        return 0.0
    datalike = sum(1 for v in non_empty if is_numeric(v) or is_date_like(v))
    return datalike / len(non_empty)


def _ffill_header_row(grid: list[list[str]], r: int, start: int, stop: int) -> list[str]:
    """Header labels for cols start..stop with horizontal forward-fill.

    Spanning headers (merged cells / repeated group labels) leave blanks that
    inherit the label to their left, but only within the value zone.
    """
    labels = []
    prev = ""
    for c in range(start, stop):
        v = clean_cell(grid_cell(grid, r, c))
        if v == "":
            v = prev
        labels.append(v)
        prev = v
    return labels


def _composed_keys(grid: list[list[str]], header_rows: "tuple[int, int]",
                   stub_end: int, right: int) -> list[str]:
    """Per-value-column keys composed from all header block levels."""
    level_labels = [_ffill_header_row(grid, r, stub_end, right)
                    for r in range(header_rows[0], header_rows[1] + 1)]
    keys = []
    for i in range(right - stub_end):
        parts = [lv[i] for lv in level_labels if lv[i] != ""]
        keys.append(" / ".join(parts))
    return keys


def _key_ambiguity(keys: list[str]) -> int:
    """How badly the composed column keys fail to identify the value columns:
    number of blank keys + number of duplicate keys. 0 = fully unambiguous."""
    if not keys:
        return 1
    blanks = sum(1 for k in keys if k == "")
    dups = len(keys) - len(set(keys))
    return blanks + dups


_ID_HINT = re.compile(r"(?i)(\bid\b|\bno\.?\b|\bindex\b|\bcode\b|\bkey\b|\bseq\b"
                      r"|id$|번호|코드|일련|연도|년도)")


def _col_profile(grid: list[list[str]], top: int, bottom: int, c: int,
                 *, max_sample: int = 1000) -> dict:
    """Type/uniqueness profile of a column over the data body (sampled)."""
    vals = []
    for r in range(top, bottom + 1):
        v = str(grid_cell(grid, r, c)).strip()
        if not is_empty(v):
            vals.append(v)
        if len(vals) >= max_sample:
            break
    n = len(vals)
    if n == 0:
        return {"n": 0, "datalike": 0.0, "date": 0.0, "int_like": False,
                "unique": 0.0, "seqish": False, "yearish": False}

    numeric = [v for v in vals if is_numeric(v)]
    dates = [v for v in vals if not is_numeric(v) and is_date_like(v)]
    ints = []
    for v in numeric:
        stripped = re.sub(r"^['\"\$₩]|[원만USD%\s]+$", "", v).replace(",", "")
        if re.fullmatch(r"[+-]?\d+", stripped):
            ints.append(int(stripped))
    int_like = len(numeric) > 0 and len(ints) == len(numeric)
    unique = len(set(vals)) / n
    seqish = False
    if int_like and ints:
        span = max(ints) - min(ints) + 1
        seqish = span <= max(len(ints), 1) * 1.5
    yearish = (int_like and ints
               and sum(1 for i in ints if 1900 <= i <= 2100) / len(ints) >= 0.8)
    return {"n": n,
            "datalike": (len(numeric) + len(dates)) / n,
            "date": len(dates) / n,
            "int_like": int_like,
            "unique": unique,
            "seqish": seqish,
            "yearish": yearish}


_MONTH_QUARTER = re.compile(
    r"(?i)^\s*("
    r"jan(uary)?|feb(ruary)?|mar(ch)?|apr(il)?|may|jun(e)?|jul(y)?|aug(ust)?"
    r"|sep(t)?(ember)?|oct(ober)?|nov(ember)?|dec(ember)?"
    r"|q[1-4]|[1-4]\s*분기|1[0-2]?월|[1-9]월"
    r")\s*$")


def _year_int(val: str) -> "int | None":
    """Parse a bare 4-digit year label (1900-2100), else None."""
    v = str(val).strip().replace(",", "")
    if re.fullmatch(r"\d{4}", v):
        y = int(v)
        if 1900 <= y <= 2100:
            return y
    return None


def _is_label_like(val: str) -> bool:
    """A cell that reads as a column label rather than a measurement:
    a bare year, a month/quarter name, or a 'wk3'/'week 3'-style token."""
    v = str(val).strip()
    if not v:
        return False
    if _year_int(v) is not None:
        return True
    if _MONTH_QUARTER.match(v):
        return True
    if re.fullmatch(r"(?i)(wk|week|w|day|d|t|period|p)\s*_?\d{1,3}", v):
        return True
    return False


def _has_header(grid: list[list[str]], h: int, top: int,
                left: int, right: int, bottom: int) -> bool:
    """Robustly decide whether row ``h`` is a header or the first data row.

    A row is a HEADER when it labels the columns below it. Two independent
    signals, either sufficient:
      (1) label-over-data: at least one column where the candidate cell is
          text but the body of that column is numeric/date (the classic
          "name over numbers").
      (2) label tokens: the candidate row carries year / month / quarter /
          'wk N' labels over columns whose body is NOT itself that kind of
          label (pivot/crosstab headers like ``country | 1999 | 2000``).
    Otherwise, if the candidate row's per-column types agree with the body's
    dominant types (row h is drawn from the same value pool), it is DATA and
    the table is headerless.
    """
    body_rows = list(range(h + 1, min(h + 60, bottom + 1)))
    if not body_rows:
        return True  # nothing below → treat the single row as a header

    def body_type(c: int) -> str:
        counts: Counter = Counter()
        for r in body_rows:
            counts[cell_type(grid_cell(grid, r, c))] += 1
        counts.pop("empty", None)
        return counts.most_common(1)[0][0] if counts else "empty"

    label_over_data = 0
    label_tokens = 0
    agree = 0
    considered = 0
    for c in range(left, right):
        cell = grid_cell(grid, h, c)
        if is_empty(cell):
            continue
        considered += 1
        htype = cell_type(cell)
        btype = body_type(c)
        if htype == "text" and btype in ("numeric", "date"):
            label_over_data += 1
        if _is_label_like(cell):
            # body of this column is a label pool too? then it's just data
            body_labels = sum(1 for r in body_rows
                              if _is_label_like(grid_cell(grid, r, c)))
            body_nonempty = sum(1 for r in body_rows
                                if not is_empty(grid_cell(grid, r, c)))
            if body_nonempty == 0 or body_labels / body_nonempty < 0.5:
                label_tokens += 1
        if htype == btype:
            agree += 1

    if considered == 0:
        return False
    if label_over_data >= 1 or label_tokens >= 1:
        return True
    # No label signal: headerless iff row h matches the body's type pattern.
    return (agree / considered) < 0.85


def _detect_stub_end(grid: list[list[str]], top: int, bottom: int,
                     left: int, right: int,
                     header_span: "tuple[int, int] | None" = None) -> int:
    """Stub = leading label block: text columns plus key-like leading columns.

    A column joins the stub prefix if it is:
      - text-dominant (labels), or
      - date-dominant AND the very first column (a timestamp index), or
      - a key-like numeric column: all-integer, mostly unique, and either
        sequential (row ids, ages, years) or year-valued or named like an id.
    The walk stops at the first measurement-like column.
    """
    if right - left <= 1:
        return left

    stub_end = left
    for c in range(left, right):
        p = _col_profile(grid, top, bottom, c)
        if p["n"] == 0:
            break  # fully empty column ends the label block
        if p["datalike"] < 0.55:
            stub_end = c + 1  # text labels
            continue
        if p["date"] >= 0.55:
            if c == left:
                stub_end = c + 1  # leading timestamp index
                continue
            break  # date column after labels = data
        # numeric column: only key-like ones belong to the stub
        hint = False
        if header_span is not None:
            for r in range(header_span[0], header_span[1] + 1):
                if _ID_HINT.search(str(grid_cell(grid, r, c))):
                    hint = True
                    break
        if (p["int_like"] and (hint or (p["unique"] >= 0.85 and p["seqish"]))) \
                or p["yearish"]:
            stub_end = c + 1
            continue
        break

    return min(stub_end, right - 1)


def build_table_model(
    grid: list[list[str]],
    region: TableRegion,
    *,
    max_header_rows: int = 5,
) -> TableModel:
    """Upgrade a coarse TableRegion into a structural TableModel.

    Decides what the region cannot express:
      1. headerless tables (region.header row that is actually a data row)
      2. multi-row header blocks — extended downward AND upward while each
         added row strictly reduces column-key ambiguity (spanning headers
         leave duplicate/blank keys until every level is included)
      3. units rows glued under the header (absorbed only when they would
         otherwise pollute the data block)
      4. stub columns (leading label/id columns before the value block)
    """
    h, top, left, bottom, right = region
    bottom = min(bottom, len(grid) - 1)
    right = min(right, grid_cols(grid))

    if h < 0 or not _has_header(grid, h, top, left, right, bottom):
        header_span = None
        data_top = min(h, top) if h >= 0 else top
    else:
        header_span, data_top = _resolve_header_block(
            grid, h, top, left, right, bottom, max_header_rows)

    # A named header column is part of the table even if its data is sparse or
    # empty (trailing 'wk NN' week columns, an empty '비고' column). Grow the
    # column span outward to cover contiguously-named header columns.
    if header_span is not None:
        left, right = _header_column_span(
            grid, header_span, left, right, data_top, bottom)

    data_top = min(data_top, bottom)
    data_bottom = _trim_trailing_summary(grid, data_top, bottom, left, right)
    stub_end = _detect_stub_end(grid, data_top, data_bottom, left, right, header_span)

    return TableModel(header_rows=header_span, stub_end=stub_end,
                      top=data_top, left=left, bottom=data_bottom, right=right)


def _resolve_header_block(grid, h, top, left, right, bottom, max_header_rows):
    """Grow a single detected header row into the full header block and return
    ``(header_span, data_top)``. See build_table_model for the rationale."""
    provisional_stub = _detect_stub_end(grid, min(h + 1, bottom), bottom, left, right)

    def ambiguity(h_start: int, h_end: int) -> int:
        return _key_ambiguity(
            _composed_keys(grid, (h_start, h_end), provisional_stub, right))

    def value_zone_empty(r: int) -> bool:
        return all(is_empty(grid_cell(grid, r, c))
                   for c in range(provisional_stub, right))

    # The picked header row may carry only a side caption (e.g. a quarter label
    # in the stub) with nothing over the value columns — advance to the first
    # row that actually labels the value zone.
    h_start = h
    advanced = 0
    while value_zone_empty(h_start) and h_start + 1 < bottom and advanced < 5:
        h_start += 1
        advanced += 1
    h_end = h_start
    score = ambiguity(h_start, h_end)

    # Extend downward while it strictly reduces column-key ambiguity and the
    # candidate row still looks like labels, not data.
    while (score > 0
           and h_end - h_start + 1 < max_header_rows
           and h_end + 1 < bottom
           and row_fill_rate(grid[h_end + 1] if h_end + 1 < len(grid) else []) > 0
           and _row_datalike_frac(grid, h_end + 1, provisional_stub, right) < 0.5):
        new_score = ambiguity(h_start, h_end + 1)
        if new_score >= score:
            break
        h_end += 1
        score = new_score

    # Extend upward (spanning group rows / year rows above the header).
    while (score > 0
           and h_end - h_start + 1 < max_header_rows
           and h_start - 1 >= 0
           and row_fill_rate(grid[h_start - 1]) > 0
           and not has_summary_keyword(grid[h_start - 1])):
        new_score = ambiguity(h_start - 1, h_end)
        if new_score >= score:
            break
        h_start -= 1
        score = new_score

    # Drop leading header rows that say nothing about the value zone.
    while h_start < h_end and value_zone_empty(h_start):
        h_start += 1

    # Units row glued under the header: absorb it only when it would otherwise
    # fall inside the data block.
    if (h_end + 1 >= top
            and h_end - h_start + 1 < max_header_rows
            and h_end + 1 < bottom
            and row_fill_rate(grid[h_end + 1] if h_end + 1 < len(grid) else []) >= 0.5
            and _row_datalike_frac(grid, h_end + 1, provisional_stub, right) <= 0.2):
        below = [_row_datalike_frac(grid, r, provisional_stub, right)
                 for r in range(h_end + 2, min(h_end + 8, bottom + 1))]
        if below and sum(below) / len(below) >= 0.6:
            h_end += 1

    return (h_start, h_end), max(h_end + 1, top)


def _header_column_span(grid, header_span, left, right, top, bottom
                        ) -> "tuple[int, int]":
    """Widen ``(left, right)`` to cover columns that are named in the header
    block and contiguously adjacent to the current span.

    A column counts as 'named' if any row of the header block has a non-empty
    cell there. Extension stops at the first unnamed column, so a stray note
    far to the side of the table cannot pull the boundary out to meet it.

    A named-but-dataless column IS kept when it is a genuine part of the frame
    (e.g. an empty '비고' column where the grid ends right after it), but NOT
    when it is an orphan header sitting at the mouth of a large empty region
    (a lone label in a 1000-column sheet): such a column has no data and is
    followed by an unnamed column, so it is treated as noise.
    """
    ncols = grid_cols(grid)
    hr0, hr1 = header_span

    def named(c: int) -> bool:
        return any(not is_empty(grid_cell(grid, r, c)) for r in range(hr0, hr1 + 1))

    def has_data(c: int) -> bool:
        return any(not is_empty(grid_cell(grid, r, c))
                   for r in range(top, bottom + 1))

    new_right = right
    while new_right < ncols and named(new_right):
        if (not has_data(new_right)
                and new_right + 1 < ncols and not named(new_right + 1)):
            break  # dead-end named-empty column opening an empty tail
        new_right += 1
    new_left = left
    while new_left - 1 >= 0 and named(new_left - 1):
        if (not has_data(new_left - 1)
                and new_left - 2 >= 0 and not named(new_left - 2)):
            break
        new_left -= 1
    return new_left, new_right


def _trim_trailing_summary(grid: list[list[str]], top: int, bottom: int,
                           left: int, right: int) -> int:
    """Pull ``bottom`` in past trailing total/subtotal/footnote/blank rows.

    Only trims the OUTER frame: it stops at the first row from the bottom that
    is genuine data, so interior subtotals are preserved. A row is trimmable
    when it is empty, carries a summary keyword (합계/total/평균/※/자료…), or is
    a lone label with an otherwise-empty value zone.
    """
    b = bottom
    while b > top:
        raw = [grid_cell(grid, b, c) for c in range(left, right)]
        if all(is_empty(v) for v in raw):
            b -= 1
            continue
        if has_summary_keyword(raw):
            b -= 1
            continue
        break
    return b


def model_leaf_header(grid: list[list[str]], model: TableModel) -> int:
    """The single row index PyNorma reports as *the* header row.

    For a multi-row header block this is the leaf-name row: the one carrying
    the most distinct labels over the value zone (group/units rows repeat or
    blank out). Returns -1 for headerless tables.
    """
    if model.header_rows is None:
        return -1
    best_row, best_distinct = model.header_rows[0], -1
    for r in range(model.header_rows[0], model.header_rows[1] + 1):
        vals = {clean_cell(grid_cell(grid, r, c))
                for c in range(model.stub_end, model.right)
                if not is_empty(grid_cell(grid, r, c))}
        # >= keeps the lowest (closest-to-data) row on ties
        if len(vals) >= best_distinct:
            best_distinct, best_row = len(vals), r
    return best_row


def model_to_region(grid: list[list[str]], model: TableModel) -> TableRegion:
    """Collapse a structural TableModel back into the 5-int TableRegion the
    public API reports — carrying the model's refined header/top/bottom."""
    return TableRegion(
        header=model_leaf_header(grid, model),
        top=model.top, left=model.left,
        bottom=model.bottom, right=model.right,
    )


def model_headers(grid: list[list[str]], model: TableModel
                  ) -> tuple[list[str], list[list[str]]]:
    """Column names implied by a TableModel.

    Returns (stub_names, level_labels):
      stub_names   — one name per stub column (header cells joined vertically)
      level_labels — per header level, one ffilled label per value column;
                     headerless tables get a single synthetic level.
    """
    stub_names = []
    for c in range(model.left, model.stub_end):
        if model.header_rows is not None:
            parts = []
            for r in range(model.header_rows[0], model.header_rows[1] + 1):
                v = clean_cell(grid_cell(grid, r, c))
                if v and (not parts or parts[-1] != v):
                    parts.append(v)
            name = " ".join(parts)
        else:
            name = ""
        stub_names.append(name if name else f"Col_{c - model.left}")

    if model.header_rows is not None:
        level_labels = [_ffill_header_row(grid, r, model.stub_end, model.right)
                        for r in range(model.header_rows[0], model.header_rows[1] + 1)]
    else:
        level_labels = [[f"Col_{c - model.left}"
                         for c in range(model.stub_end, model.right)]]
    return stub_names, level_labels


def to_long(
    grid: list[list[str]],
    model: TableModel,
    *,
    value_name: str = "value",
    dropna: bool = True,
    ffill_stub: bool = True,
    skip_summary: bool = True,
) -> tuple[list[str], list[list[str]]]:
    """Convert a modeled table region into long form.

    Each output row = [*stub_values, *header_level_labels, value] — one row
    per non-empty value cell. Deterministic given the model: this is where
    "any tabular data → long form" actually happens.

    Returns (columns, rows).
    """
    stub_names, level_labels = model_headers(grid, model)
    n_levels = len(level_labels)
    level_names = (["variable"] if n_levels == 1
                   else [f"level_{j + 1}" for j in range(n_levels)])
    columns = [*stub_names, *level_names, value_name]

    rows: list[list[str]] = []
    stub_memory = [""] * (model.stub_end - model.left)

    for r in range(model.top, min(model.bottom + 1, len(grid))):
        raw_row = [grid_cell(grid, r, c) for c in range(model.left, model.right)]
        cleaned = [clean_cell(v) for v in raw_row]
        if all(v == "" for v in cleaned):
            continue
        if skip_summary and has_summary_keyword(raw_row):
            continue

        stub_vals = []
        for i, c in enumerate(range(model.left, model.stub_end)):
            v = cleaned[c - model.left]
            if v == "" and ffill_stub:
                v = stub_memory[i]
            else:
                stub_memory[i] = v
            stub_vals.append(v)

        for i, c in enumerate(range(model.stub_end, model.right)):
            v = cleaned[c - model.left]
            if dropna and v == "":
                continue
            rows.append([*stub_vals, *[lv[i] for lv in level_labels], v])

    return columns, rows


def clean_region_model(grid: list[list[str]], model: TableModel) -> list[list[str]]:
    """Model-aware variant of clean_region: wide output with composed,
    disambiguated column names (multi-row headers joined with ' / ')."""
    stub_names, level_labels = model_headers(grid, model)
    value_keys = []
    for i in range(model.right - model.stub_end):
        parts = [lv[i] for lv in level_labels if lv[i] != ""]
        key = " / ".join(parts)
        value_keys.append(key if key else f"Col_{model.stub_end - model.left + i}")

    headers = [*stub_names, *value_keys]
    result = [headers]
    seen = set()

    for r in range(model.top, min(model.bottom + 1, len(grid))):
        raw_row = [grid_cell(grid, r, c) for c in range(model.left, model.right)]
        row = [clean_cell(v) for v in raw_row]
        if all(v == "" for v in row):
            continue
        if has_summary_keyword(raw_row):
            continue
        key = tuple(row)
        if key in seen:
            continue
        seen.add(key)
        result.append(row)

    return result


# ═══════════════════════════════════════════
# Internal Quality Score (GT-free)
# ═══════════════════════════════════════════

def quality_score(grid: list[list[str]], region: TableRegion) -> float:
    """Score a TableRegion without ground truth.

    Higher = cleaner, more consistent table. Range: 0.0 ~ 1.0.
    """
    h, top, left, bottom, right = region
    nrows = bottom - top + 1
    ncols = right - left
    if nrows <= 0 or ncols <= 0:
        return 0.0

    # 1. Type consistency per column (0~1)
    col_scores = []
    for c in range(left, right):
        types = [cell_type(grid_cell(grid, r, c))
                 for r in range(top, min(bottom + 1, len(grid)))]
        non_empty = [t for t in types if t != "empty"]
        if non_empty:
            most = Counter(non_empty).most_common(1)[0][1]
            col_scores.append(most / len(non_empty))
        else:
            col_scores.append(0.5)
    type_consistency = sum(col_scores) / len(col_scores) if col_scores else 0.0

    # 2. Fill uniformity (inverse of variance, 0~1)
    fills = [row_fill_count([grid_cell(grid, r, c) for c in range(left, right)])
             for r in range(top, min(bottom + 1, len(grid)))]
    if len(fills) > 1:
        mean_f = sum(fills) / len(fills)
        var_f = sum((f - mean_f) ** 2 for f in fills) / len(fills)
        max_var = (ncols ** 2) / 4  # max possible variance
        fill_uniformity = 1.0 - min(var_f / max(max_var, 1), 1.0)
    else:
        fill_uniformity = 1.0

    # 3. Header confidence (text ratio in header row, 0~1)
    if h >= 0 and h < len(grid):
        hdr_cells = [grid_cell(grid, h, c) for c in range(left, right)]
        non_empty_hdr = [c for c in hdr_cells if not is_empty(c)]
        if non_empty_hdr:
            text_count = sum(1 for c in non_empty_hdr if not is_numeric(c))
            header_confidence = text_count / len(non_empty_hdr)
        else:
            header_confidence = 0.0
    else:
        header_confidence = 0.5

    # 4. Boundary sharpness (density contrast at edges, 0~1)
    data_density = sum(fills) / (nrows * ncols) if (nrows * ncols) > 0 else 0
    edge_densities = []
    # row above top
    if top > 0:
        above = row_fill_count([grid_cell(grid, top - 1, c) for c in range(left, right)])
        edge_densities.append(above / max(ncols, 1))
    # row below bottom
    if bottom + 1 < len(grid):
        below = row_fill_count([grid_cell(grid, bottom + 1, c) for c in range(left, right)])
        edge_densities.append(below / max(ncols, 1))
    if edge_densities:
        boundary_sharpness = max(0, data_density - sum(edge_densities) / len(edge_densities))
    else:
        boundary_sharpness = data_density  # no edges = grid boundary = sharp

    # 5. Coverage (how much of the grid does this region use)
    total_cells = len(grid) * grid_cols(grid) if grid else 1
    region_cells = nrows * ncols
    coverage = min(region_cells / max(total_cells, 1), 1.0)

    # 6. Size factor (penalize very small regions that look artificially clean)
    total_rows = len(grid)
    size_ratio = nrows / max(total_rows, 1)
    # Small regions (< 10% of grid) get penalized; bigger regions get full credit
    size_factor = min(size_ratio * 5, 1.0)  # reaches 1.0 at 20% of grid

    score = (0.25 * type_consistency
             + 0.15 * fill_uniformity
             + 0.20 * header_confidence
             + 0.15 * min(boundary_sharpness * 2, 1.0)  # scale up
             + 0.15 * coverage
             + 0.10 * size_factor)

    return round(score, 4)


# ═══════════════════════════════════════════
# Metrics (for benchmark, needs GT)
# ═══════════════════════════════════════════

@dataclass
class Scores:
    header_accuracy: float = 0.0
    boundary_iou: float = 0.0
    noise_removal: float = 0.0
    data_preservation: float = 0.0
    cleaning_quality: float = 0.0

    @property
    def avg(self) -> float:
        return (self.header_accuracy + self.boundary_iou +
                self.noise_removal + self.data_preservation +
                self.cleaning_quality) / 5


def region_to_row_set(region: TableRegion) -> set[int]:
    """Convert a region to a set of row indices (data rows only)."""
    return set(range(region.top, region.bottom + 1))


def compute_scores(
    pred_regions: list[TableRegion],
    gt_regions: list[TableRegion],
    grid: list[list[str]],
) -> Scores:
    """Compute benchmark scores: predicted regions vs ground truth regions."""
    s = Scores()

    if not gt_regions:
        s.header_accuracy = 1.0 if not pred_regions else 0.0
        s.boundary_iou = 1.0 if not pred_regions else 0.0
        s.noise_removal = 1.0
        s.data_preservation = 1.0
        s.cleaning_quality = 0.5
        return s

    # Match predicted to GT by largest overlap (greedy)
    gt_used = set()
    matches: list[tuple[TableRegion, TableRegion]] = []

    for pred in pred_regions:
        best_iou, best_gt_idx = -1, -1
        pred_rows = region_to_row_set(pred)
        for gi, gt in enumerate(gt_regions):
            if gi in gt_used:
                continue
            gt_rows = region_to_row_set(gt)
            inter = len(pred_rows & gt_rows)
            union = len(pred_rows | gt_rows)
            iou = inter / union if union > 0 else 0
            if iou > best_iou:
                best_iou, best_gt_idx = iou, gi
        if best_gt_idx >= 0:
            gt_used.add(best_gt_idx)
            matches.append((pred, gt_regions[best_gt_idx]))

    if not matches:
        return s

    # 1. Header accuracy (avg over matched pairs)
    h_scores = [1.0 / (1.0 + abs(p.header - g.header)) for p, g in matches]
    s.header_accuracy = sum(h_scores) / len(h_scores)

    # 2. Boundary IoU (avg over matched pairs)
    ious = []
    for pred, gt in matches:
        pred_cells = {(r, c) for r in range(pred.top, pred.bottom + 1) for c in range(pred.left, pred.right)}
        gt_cells = {(r, c) for r in range(gt.top, gt.bottom + 1) for c in range(gt.left, gt.right)}
        inter = len(pred_cells & gt_cells)
        union = len(pred_cells | gt_cells)
        ious.append(inter / union if union > 0 else 0)
    s.boundary_iou = sum(ious) / len(ious)

    # 3. Noise removal: rows outside all GT regions that are also outside all pred regions
    all_rows = set(range(len(grid)))
    gt_data_rows = set()
    for gt in gt_regions:
        gt_data_rows |= region_to_row_set(gt)
        if gt.header >= 0:
            gt_data_rows.add(gt.header)
    noise_rows = all_rows - gt_data_rows
    pred_data_rows = set()
    for pred in pred_regions:
        pred_data_rows |= region_to_row_set(pred)
        if pred.header >= 0:
            pred_data_rows.add(pred.header)
    pred_excluded = all_rows - pred_data_rows
    if noise_rows:
        s.noise_removal = len(noise_rows & pred_excluded) / len(noise_rows)
    else:
        s.noise_removal = 1.0

    # 4. Data preservation: GT data rows that are in pred regions
    if gt_data_rows:
        s.data_preservation = len(gt_data_rows & pred_data_rows) / len(gt_data_rows)
    else:
        s.data_preservation = 1.0

    # 5. Cleaning quality: type consistency of cleaned output
    cq_scores = []
    for pred, _ in matches:
        cleaned = clean_region(grid, pred)
        if len(cleaned) > 1:
            data = cleaned[1:]
            ncols = len(cleaned[0])
            for col in range(ncols):
                types = [cell_type(r[col]) for r in data if col < len(r)]
                non_empty = [t for t in types if t != "empty"]
                if non_empty:
                    cq_scores.append(Counter(non_empty).most_common(1)[0][1] / len(non_empty))
                else:
                    cq_scores.append(1.0)
    s.cleaning_quality = sum(cq_scores) / len(cq_scores) if cq_scores else 0.5

    # Penalty for wrong table count
    count_penalty = abs(len(pred_regions) - len(gt_regions)) * 0.05
    s.boundary_iou = max(0, s.boundary_iou - count_penalty)

    return s
